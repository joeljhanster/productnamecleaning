import pandas as pd
import numpy as np
import re
from openpyxl import *
import copy
import ast

# Common Units for Volume
unwanted = {'FLXBN', 'ALSCN', 'FXBPN', 'PLBTN', 'GLBTN', 'ALCNN', 'ALSCN'}

# Path to Excel
FILEPATH = '/Users/jiehan/Desktop/HiMart/Data Cleaning Code/Masterlist160520 - Joel.xlsx'
SHEET = 'vegetables'
BRANDPATH = 'brand_list.txt'
QUANTITYPATH = 'quantity_lst.txt'

# Read Excel Sheet
df = pd.read_excel(FILEPATH,sheet_name=SHEET)
array = np.array(df)

# Number of columns and rows
columns = len(array[0,:])
rows = len(array[:,0])

# Find Volume from Product Name
def changeVolume(ws,i):

    vol_cell = ws.cell(i+2,6)   # get specific volume cells
    proc_cell = ws.cell(i+2,8)      # get specific processing cells

    product_string = proc_cell.value

    if (proc_cell.value != None and vol_cell.value == None):
        match = re.search(r'((\d\.?)+(ml|oz|l|g|kg|cl|mm|cm|ft| kg))', product_string, re.I)

        if match:
            # Find volume in string
            volume = match.group(1)

            # Fill in Volumn column
            vol_cell.value = volume.lower()

            # Remove volume from original string
            proc_cell.value = product_string.replace(volume,'')

        # Remove unwanted
        for element in unwanted:
            if element in proc_cell.value:
                substrings = proc_cell.value.split()
                new_string = copy.deepcopy(substrings)
                for substring in substrings:
                    if element in substring:
                        new_string.remove(substring)
                        break
                proc_cell.value = ' '.join(new_string)
                break

    print('Volume added')

# Find Brand from Product Name
def changeBrand(ws,i):

    brand_cell = ws.cell(i+2,4)     # get specific brand cells
    proc_cell = ws.cell(i+2,8)      # get specific processing cells
    product_cell = ws.cell(i+2,5)   # get specific product name cells

    product_string = proc_cell.value

    if (proc_cell.value != None):
        if (brand_cell.value == None and product_cell.value == None):
            brand_name, new_string = inputBrand(product_string)

            brand_cell.value = brand_name
            proc_cell.value = new_string

            if (brand_name == ""):  # No Brand
                add_brand = raw_input("Do you want to add brand name?(Y/N) ").lower()
                if (add_brand == 'y'):
                    brand_name = raw_input("Write new brand name: ")
                    brand_cell.value = brand_name

    print('Brand added')

def changeQuantity(ws, i):

    brand_cell = ws.cell(i+2,4)     # get specific brand cells
    proc_cell = ws.cell(i+2,8)      # get specific processing cells
    product_cell = ws.cell(i+2,5)   # get specific product name cells
    quantity_cell = ws.cell(i+2,7)  # get specific quantity cells

    product_string = proc_cell.value

    if (proc_cell.value != None or quantity_cell.value != None):
        # CLEAN UP REMAINING NAME: Remove Quantity from Product Name
        contains_digit = any(map(str.isdigit, str(product_string)))
        if (contains_digit):
            print("This is the new string: {0}".format(product_string))
            quantity, product_string = inputQuantity(product_string)
            quantity_cell.value = quantity
            proc_cell.value = product_string

    print('Quantity added')

# Find Product Name
def changeProduct(ws, i):

    brand_cell = ws.cell(i+2,4)     # get specific brand cells
    proc_cell = ws.cell(i+2,8)      # get specific processing cells
    product_cell = ws.cell(i+2,5)   # get specific product name cells
    quantity_cell = ws.cell(i+2,7)  # get specific quantity cells

    product_string = proc_cell.value

    if (proc_cell.value != None):

        # No product name
        if (product_cell.value == None):
            product_string = product_string.title()
            print("This is the new string: {0}".format(product_string))
            add_product = raw_input("Do you want to add product name?(Y/N) ").lower()
            if (add_product == 'y'):
                change_product = raw_input("Do you want to use this product name (Y) or change it (N)?").lower()
                if (change_product == 'y'):
                    product_cell.value = product_string
                    proc_cell.value = None
                elif (change_product == 'n'):
                    product_cell.value = raw_input("Write new product name: ")
                    proc_cell.value = None

    print('Product added')


def inputQuantity(string):
    f = open(QUANTITYPATH, "r")
    quantity_dict = ast.literal_eval(f.readline())
    f.close()

    string = string.lower()

    while(True):
        if len(quantity_dict) == 0:
            add_quantity = raw_input("Do you want to add quantity?(Y/N) ").lower()
            if add_quantity == 'y':
                quantity_value = raw_input("What is the quantity? ")
                quantity_key = raw_input("Add quantity abbreviation (based on product name): ").lower()
                quantity_dict[quantity_key] = quantity_value

                f = open(QUANTITYPATH, "w")
                f.write(str(quantity_dict))
                f.close()

                new_string = string.replace(quantity_key, '')
                return quantity_value, new_string
            elif add_quantity == 'n':
                return "", string
            
        if len(quantity_dict) > 0:
            for key in sorted(quantity_dict, key=len, reverse=True):
                value = quantity_dict[key]
                if key in string:
                    quantity = value
                    new_string = string.replace(key, '')
                    return quantity, new_string

            add_quantity = raw_input("Do you want to add quantity?(Y/N) ").lower()
            if add_quantity == 'y':
                quantity_value = raw_input("What is the quantity? ")
                quantity_key = raw_input("Add quantity abbreviation (based on product name): ").lower()
                quantity_dict[quantity_key] = quantity_value

                f = open(QUANTITYPATH, "w")
                f.write(str(quantity_dict))
                f.close()

                new_string = string.replace(quantity_key, '')
                return quantity_value, new_string
            elif add_quantity == 'n':
                return "", string


def inputBrand(string):
    f = open(BRANDPATH, "r")
    brand_dict = ast.literal_eval(f.readline())
    f.close()

    string = string.lower()

    while (True):
        if len(brand_dict) > 0:

            test_string = ''
            length = len(string.split())
            if length > 3:
                length = 3

            for i in range(length):
                if i == 0:
                    test_string = string.split()[i]
                else:
                    test_string = test_string + ' ' + string.split()[i]

                for key in brand_dict:
                    if test_string.strip().lower() == key.strip().lower():

                        new_string = string.replace(key.lower(),'')
                        brand_name = brand_dict[key]

                        f = open(BRANDPATH, "w")
                        f.write(str(brand_dict))
                        f.close()

                        return (brand_name, new_string)

            for key in sorted(brand_dict, key=len, reverse=True):
                value = brand_dict[key]
                # Additional Spacings
                if (len(key.split()) > 1):
                    if key.lower() in string:
                        new_string = string.replace(key.lower(),'')
                        brand_name = value

                        f = open(BRANDPATH, "w")
                        f.write(str(brand_dict))
                        f.close()

                        return (brand_name, new_string)

                # Proper Brand (brand in set) and Capitalized, e.g. ZICO => Zico
                elif key.lower() in string.split():
                    new_string = string.replace(key.lower(),'')
                    brand_name = value

                    # Write into brand_list.txt
                    f = open(BRANDPATH, "w")
                    f.write(str(brand_dict))
                    f.close()

                    return (brand_name, new_string)

                elif key.lower() in string:
                    new_string = string.replace(key.lower(),'')
                    brand_name = value

                    f = open(BRANDPATH, "w")
                    f.write(str(brand_dict))
                    f.close()

                    return (brand_name, new_string)
    
        is_brand = raw_input("String is: {0}\nIs there a brand? (Y/N)".format(string)).lower()

        # No Brand, e.g. PEACH TEA
        if (is_brand == "n"):
            print ("No Brand")
            return ("", string)

        # Abbreviation, e.g. MM => Minute Maid; Additional Spacings, e.g. A & W => A&W; Incomplete, e.g. Authentic Tea vs Authentic Tea House
        elif (is_brand == "y"):
            brand_value = raw_input("What is the brand name? ")
            brand_key = raw_input("Add brand name/abbreviation (based on product name): ").capitalize()
            brand_dict[brand_key] = brand_value

# Load Excel File
wb = load_workbook(FILEPATH)
ws = wb[SHEET]

for i in range(rows):
    # changeVolume(ws, i)
    changeBrand(ws, i)
    # changeQuantity(ws, i)
    # changeProduct(ws, i)

    proc_cell = ws.cell(i+2, 8)

    if (proc_cell.value != None):
        proc_cell.value = proc_cell.value.title()

    print("Done with line {}".format(i+2))

    # final_name = ''
    # for j in range(4,8):
    #     cell = ws.cell(i+2,j)
    #     if cell.value != None:
    #         final_name = final_name + ' ' + cell.value
    # print("Final product name for line{0}: {1}".format(i+2,final_name))
    # if (ws.cell(i+2,8).value != None):
    #     print ("Processing is left with: {}".format(ws.cell(i+2,8).value))

    # Save excel file after every 30
    if (i % 50 == 0):
        wb.save(FILEPATH)
        print('Saving')

wb.save(FILEPATH)
print('Saving')